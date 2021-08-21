
#!/usr/bin/env python2.7
# -*- coding: utf-8 -*-
import matplotlib.pyplot as plt
import numpy as np
from scipy.ndimage import gaussian_filter
from scipy.signal import find_peaks
# from data_processing.read_data import read_file
from openpyxl import Workbook, load_workbook


def read_file(path: str) -> tuple:
    """recives a path of a excel file and return a tuple of two lists, 
    corresponding to Edi (uV) and Pes"""
    wb = Workbook()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Resumen para análisis"]

    output_edi = []
    output_pes = []

    for row in ws.rows:
        output_edi.append(row[0].value)
        output_pes.append(row[1].value)

    output_edi.pop(0)
    output_pes.pop(0)

    #Sometimes the data read Nones at the end,
    #there should be no Nones in the middle
 
    for i in range(len(output_edi)):
        if output_edi[i] == None:
            output_edi = output_edi[:i]
            break

    for i in range(len(output_pes)):
        if output_pes[i] == None:
            output_pes = output_pes[:i]
            break

    return (output_edi, output_pes)

def get_signal_peaks(data, first_dp, last_dp, min_max, max_min, dist=150) -> tuple:
    """recives a list with a signal, and tries to find its inflection point,
    returning it as a tuple (xpos,ypos). Not finished
    """

    # xs corresponds to x coordinates and ys to the signal values
    xs, ys = np.arange(last_dp-first_dp), np.array(data[first_dp:last_dp])

    # Signal "energy"
    # yss = np.square(ys)

    # inverted signal
    ysi = -1*ys

    # smooth out noise
    # smoothed = gaussian_filter(ys, 3.)

    # inverted smoothed signal
    # smoothedi = -1*smoothed

    # positions for positive peaks (local maxima)
    peaks, _ = find_peaks(ys, height=min_max, distance=dist)
    # positions for negative peaks (local minima)
    antipeaks, _ = find_peaks(ysi, height=(max_min, 0), distance=dist)

    # positions for positive peaks (local maxima)
    # peaks, _ = find_peaks(smoothed, height = min_max, distance = dist)
    # positions for negative peaks (local minima)
    # antipeaks, _ = find_peaks(smoothedi, height = (max_min, 0), distance = dist)
    return (peaks, antipeaks)


if __name__ == '__main__':
    # positions for positive peaks (local maxima)
    path = 'G:/My Drive/Work/TRabajo medico invierno 2021/data/01_ PVE 1_Importacion de datos filtrados_con graficos.xlsm'
    edi, pes = read_file(path)
    
    signl = 1
    if signl:
        signl = edi
    else:
        signl = pes
    xs, ys = np.arange(len(signl)), np.array(signl)
    peaks, antipeaks = get_signal_peaks(signl, 0, len(signl), 1.8, -1)
    # get_signal_peaks(pes, 23500, 25000, 1.8, -1)
    # smooth out noise
    smoothies = []
    for x in [0,9,20, 100]:
        smoothies.append(gaussian_filter(ys, x))
    
    # Ploting sheet
    for i, x in enumerate([0,9,20, 100]):
        plt.plot(xs, smoothies[i], '-')
    plt.plot(xs, ys, '.')
    # plt.plot(xs[peaks], ys[peaks], 'o')
    # plt.plot(xs[antipeaks], ys[antipeaks], '+')
    plt.plot(xs[peaks], smoothies[1][peaks], 'o')
    plt.plot(xs[antipeaks], smoothies[1][antipeaks], '+')
    plt.show()

