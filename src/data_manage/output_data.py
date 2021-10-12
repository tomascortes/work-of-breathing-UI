from openpyxl import Workbook, load_workbook
import datetime
from pathlib import Path
import numpy as np
import os


def create_excel(integ_data_edi, integ_data_pes, peaks_edi, anti_peaks_edi, f_name="") -> tuple:
    """recives a list of lists with shape:
    [integral_value, start_integral, end_integral]
    and returns and creates an excel file with the data in the
    output path. Creates output directory if it doesn't exist."""

    # Create workbook
    wb = Workbook()
    ws = wb.get_sheet_by_name('Sheet')
    ws.title = "results"

    # Write colum names
    colum_names = [
        "n cycle",
        "integral value pes",
        "integral value edi",
        "start pes ",
        "start edi ",
        "point 75% ",
        "t start pes->75%",
        "t start pes-> start edi",
        "t start edi -> 75%",
        "t start edi -> peak edi",
        "t start edi -> end edi",
        "edi Amplitude",
        "pes Amplitude"
    ]
    ws.append(colum_names)

    # Write data
    for cycle in range(len(integ_data_edi)):

        start_pes = integ_data_edi[cycle][1]
        t_75 = integ_data_edi[cycle][2]

        def abs_difference(list_value): return abs(list_value - start_pes)
        start_edi = min(anti_peaks_edi, key=abs_difference)
        index_s_edi = np.where(anti_peaks_edi == start_edi)[0][0]

        index_peak_edi = next_edi_peak(peaks_edi, start_pes)
        peak_edi = peaks_edi[index_peak_edi]

        if len(anti_peaks_edi) > cycle + 1:
            minus_a = anti_peaks_edi[index_s_edi + 1] - start_edi
            start_to_end_edi = (minus_a)/100
        else:
            start_to_end_edi = "ultimo peak"

        ap_aux = [
            cycle,  # Number of cycle
            integ_data_pes[cycle][0],  # value integ pes
            integ_data_edi[cycle][0],  # value integ edi
            start_pes,  # start pes
            start_edi,  # start edi
            t_75,  # end pes and 75%
            (t_75 - start_pes)/100,  # margin pes to 75%
            (start_edi - start_pes)/100,  # margin start edi to start pes
            (t_75 - start_edi)/100,  # margin edi to 75%
            (peak_edi - start_edi)/100,  # margin edi to peak
            start_to_end_edi,  # margin edi complete
            integ_data_edi[cycle][3],  # Amplitude edi
            integ_data_pes[cycle][3],  # Amplitude pes
        ]

        ws.append(ap_aux)

    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    f_name = os.path.splitext(f_name)[0]
    file_name = f_name + "_" + timestamp + ".xlsx"
    output_dir_path = "calculated_data"
    # Create output folder if it doesnt exists
    Path(output_dir_path).mkdir(parents=True, exist_ok=True)
    # expand width form columns
    for col in ws.columns:
        column = col[0].column_letter   # Get the column name
        ws.column_dimensions[column].width = 20
    # Save excel file
    wb.save(output_dir_path + "/" + file_name)


def next_edi_peak(peaks_edi, start_edi):
    for i in range(len(peaks_edi)):
        if peaks_edi[i] > start_edi:
            return i

def create_excel_without_edi( integ_data_pes, f_name=""):
    """Create excel just for the pes data."""

    # Create workbook
    wb = Workbook()
    ws = wb.get_sheet_by_name('Sheet')
    ws.title = "results"

    # Write colum names
    colum_names = [
        "n cycle",
        "integral value pes",
        "start pes ",
        "end pes ",
        "pes Amplitude"
    ]
    ws.append(colum_names)

    # Write data
    for cycle in range(len(integ_data_pes)):
        ap_aux = [
            cycle,  # Number of cycle
            integ_data_pes[cycle][0],  # value integ pes
            integ_data_pes[cycle][1],  # start pes
            integ_data_pes[cycle][2],  # end pes
            integ_data_pes[cycle][3],  # Amplitude pes
        ]

        ws.append(ap_aux)

    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    f_name = os.path.splitext(f_name)[0]
    file_name = f_name + "_" + timestamp + ".xlsx"
    output_dir_path = "calculated_data"
    # Create output folder if it doesnt exists
    Path(output_dir_path).mkdir(parents=True, exist_ok=True)
    # expand width form columns
    for col in ws.columns:
        column = col[0].column_letter   # Get the column name
        ws.column_dimensions[column].width = 20
    # Save excel file
    wb.save(output_dir_path + "/" + file_name)
