from openpyxl import Workbook, load_workbook


def read_file(path: str) -> tuple:
    """recives a path of a excel file and return a tuple of two lists, 
    corresponding to Edi (uV) and Pes"""
    wb = Workbook()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Resumen para análisis"]

    output_edi = []
    output_pes = []

    for row in ws.rows:
        output_edi.append(row[0].value)
        output_pes.append(row[1].value)

    output_edi.pop(0)
    output_pes.pop(0)

    # Sometimes the data read Nones at the end,
    # there should be no Nones in the middle

    for i in range(len(output_edi)):
        if output_edi[i] == None:
            output_edi = output_edi[:i]
            break

    for i in range(len(output_pes)):
        if output_pes[i] == None:
            output_pes = output_pes[:i]
            break

    return (output_edi, output_pes)


if __name__ == "__main__":
    print("Executing data_manage")
    read_file("./data/01_ PVE 1_Importacion de datos filtrados_con graficos.xlsm")
