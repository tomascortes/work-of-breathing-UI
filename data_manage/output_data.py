from openpyxl import Workbook, load_workbook
import datetime
from pathlib import Path


def create_excel(integral_values, output_path="") -> tuple:
    """recives a list of lists with shape:
    [integral_value, start_integral, end_integral]
    and returns and creates an excel file with the data in the
    output path. Creates output directory if it doesn't exist."""

    # Create workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = "results"

    # Write colum names
    colum_names = ["integral_value", "start_integral", "end_integral"]
    ws.append(colum_names)

    # Write data
    for i in range(len(integral_values)):
        ws.append(integral_values[i])

    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    file_name = "output_" + timestamp + ".xlsx"
    output_dir_path = "output_data"
    # Create output folder if it doesnt exists
    Path(output_dir_path).mkdir(parents=True, exist_ok=True)
    # Save excel file
    wb.save(output_dir_path + "/" + file_name)
